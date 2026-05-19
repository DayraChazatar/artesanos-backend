# usuarios/views.py
from rest_framework import viewsets
from rest_framework.authtoken.models import Token
from django.contrib.auth.models import User
from rest_framework.decorators import api_view, action, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from django.contrib.auth.hashers import check_password
from django.db.models import Sum
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# ── Imports de modelos ────────────────────────────────────────────────────────
from .models import Usuario, Categoria, Producto, Notificacion
from inventario.models import Kardex

# ── Imports de serializers ────────────────────────────────────────────────────
from .serializers import (
    UsuarioSerializer,
    CategoriaSerializer,
    ProductoSerializer,
    CatalogoProductoSerializer,
    KardexSerializer,
    NotificacionSerializer,
)


# ── Usuarios ──────────────────────────────────────────────────────────────────
class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer

    @action(detail=False, methods=['get'], url_path='artesanos')
    def artesanos(self, request):
        qs = Usuario.objects.filter(tipo='artesano')
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)


@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    correo   = request.data.get('correo')
    password = request.data.get('password')
    try:
        usuario = Usuario.objects.get(correo=correo)
        if check_password(password, usuario.password):
            auth_user, _ = User.objects.get_or_create(username=correo)
            auth_user.set_password(password)
            auth_user.save()
            token, _ = Token.objects.get_or_create(user=auth_user)
            return Response({
                'success': True,
                'id':      usuario.id,
                'nombre':  usuario.nombre,
                'tipo':    usuario.tipo,
                'token':   token.key,
            })
        return Response({'success': False, 'mensaje': 'Contraseña incorrecta'})
    except Usuario.DoesNotExist:
        return Response({'success': False, 'mensaje': 'Usuario no encontrado'})


# ── Categorías ────────────────────────────────────────────────────────────────
class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    permission_classes = [AllowAny]

    def create(self, request, *args, **kwargs):
        print("DATA RECIBIDA:", request.data)
        ser = CategoriaSerializer(data=request.data)
        ser.is_valid()
        print("ERRORES:", ser.errors)
        return super().create(request, *args, **kwargs)

    def get_queryset(self):
        qs = super().get_queryset()
        artesano_id = self.request.query_params.get('artesano')
        if artesano_id:
            qs = qs.filter(artesano_id=artesano_id)
        return qs


# ── Productos ─────────────────────────────────────────────────────────────────
class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    permission_classes = [AllowAny]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def create(self, request, *args, **kwargs):
        print("🔥 ENTRO AL CREATE")
        print("DATA:", request.data)
        print("FILES:", request.FILES)
        return super().create(request, *args, **kwargs)

    def get_queryset(self):
        qs = super().get_queryset()
        artesano_id = self.request.query_params.get('artesano')
        if artesano_id:
            qs = qs.filter(artesano_id=artesano_id)
        return qs


@api_view(['GET'])
@permission_classes([AllowAny])
def catalogo_productos(request):
    productos = Producto.objects.filter(cantidad__gt=0)
    serializer = CatalogoProductoSerializer(
        productos, many=True, context={'request': request}
    )
    return Response(serializer.data)


# ── Helper notificaciones ─────────────────────────────────────────────────────
def crear_notificacion(tipo, titulo, detalle, referencia_id=None, ruta=''):
    Notificacion.objects.create(
        tipo=tipo,
        titulo=titulo,
        detalle=detalle,
        referencia_id=referencia_id,
        ruta=ruta,
    )


# ── Kardex ────────────────────────────────────────────────────────────────────
class KardexViewSet(viewsets.ModelViewSet):
    queryset = Kardex.objects.all().order_by('-fecha')
    serializer_class = KardexSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        qs = super().get_queryset()
        producto_id = self.request.query_params.get('producto')
        if producto_id:
            qs = qs.filter(producto_id=producto_id)
        return qs

    def perform_create(self, serializer):
        kardex = serializer.save()
        producto = kardex.producto
        if producto.cantidad <= producto.stock_minimo:
            crear_notificacion(
                tipo='stock',
                titulo='Stock bajo',
                detalle=f'El producto "{producto.nombre}" tiene solo {producto.cantidad} unidades disponibles.',
                referencia_id=producto.id,
                ruta='/inventario',
            )


# ── Helper: queryset kardex con filtros ───────────────────────────────────────
def _kardex_filtrado(request):
    """Devuelve un queryset de Kardex aplicando artesano, desde y hasta."""
    artesano_id = request.query_params.get('artesano')
    desde       = request.query_params.get('desde')
    hasta       = request.query_params.get('hasta')
    qs = Kardex.objects.all().order_by('fecha')
    if artesano_id:
        qs = qs.filter(producto__artesano_id=artesano_id)
    if desde:
        qs = qs.filter(fecha__date__gte=desde)
    if hasta:
        qs = qs.filter(fecha__date__lte=hasta)
    return qs


# ── Helpers Excel/PDF ─────────────────────────────────────────────────────────
def estilo_excel(ws, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='B45309')
        cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4


def estilo_pdf(elements, titulo, headers, data):
    styles = getSampleStyleSheet()
    elements.append(Paragraph(f'<b>{titulo}</b>', styles['Title']))
    elements.append(Spacer(1, 12))
    tabla = Table([headers] + data)
    tabla.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#B45309')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  10),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#FEF3C7')]),
        ('GRID',          (0, 0), (-1, -1), 0.5, colors.HexColor('#D97706')),
        ('FONTSIZE',      (0, 1), (-1, -1), 9),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(tabla)


# ── Reportes Productos ────────────────────────────────────────────────────────
@api_view(['GET'])
def reporte_productos_excel(request):
    artesano_id = request.query_params.get('artesano')
    productos = Producto.objects.filter(artesano_id=artesano_id) if artesano_id else Producto.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'
    headers = ['Código', 'Lote', 'Nombre', 'Categoría', 'Precio neto', 'PVP', 'IVA (%)', 'Descuento', 'Stock', 'Stock mín.', 'Stock máx.', 'Estado']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for p in productos:
        ws.append([
            p.codigo_barra or '—', p.lote or '—', p.nombre,
            p.categoria.nombre if p.categoria else '—',
            float(p.precio_neto),
            float(p.precio_final),
            p.iva,
            f'Sí ({p.valor_descuento}%)' if p.descuento else 'No',
            p.cantidad, p.stock_minimo, p.stock_maximo,
            p.estado_stock,
        ])
    estilo_excel(ws, headers)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': 'attachment; filename="productos.xlsx"'})


@api_view(['GET'])
def reporte_productos_pdf(request):
    artesano_id = request.query_params.get('artesano')
    productos = Producto.objects.filter(artesano_id=artesano_id) if artesano_id else Producto.objects.all()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    headers = ['Código', 'Nombre', 'Categoría', 'Precio neto', 'PVP', 'IVA', 'Stock', 'Mín.', 'Máx.', 'Estado']
    data = [[
        p.codigo_barra or '—', p.nombre,
        p.categoria.nombre if p.categoria else '—',
        f'${float(p.precio_neto):,.0f}',
        f'${float(p.precio_final):,.0f}',
        f'{p.iva}%',
        str(p.cantidad), str(p.stock_minimo), str(p.stock_maximo),
        p.estado_stock,
    ] for p in productos]
    estilo_pdf(elements, 'Reporte de Productos — Pakari Shop', headers, data)
    doc.build(elements)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf',
                        headers={'Content-Disposition': 'attachment; filename="productos.pdf"'})


# ── Reportes Inventario ───────────────────────────────────────────────────────
@api_view(['GET'])
def reporte_inventario_excel(request):
    artesano_id = request.query_params.get('artesano')
    productos = Producto.objects.filter(artesano_id=artesano_id) if artesano_id else Producto.objects.all()
    wb = Workbook()

    # Hoja 1: stock por producto
    ws = wb.active
    ws.title = 'Stock'
    headers = ['Producto', 'Código', 'Stock actual', 'Stock mínimo', 'Stock máximo', 'Estado']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for p in productos:
        ws.append([p.nombre, p.codigo_barra or '—', p.cantidad, p.stock_minimo, p.stock_maximo, p.estado_stock])
    estilo_excel(ws, headers)

    # Hoja 2: movimientos (kardex)
    ws2 = wb.create_sheet(title='Movimientos')
    headers2 = ['Producto', 'Tipo', 'Subtipo', 'Cantidad', 'Stock result.', 'Origen', 'Pedido', 'Fecha', 'Nota']
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    kardex = _kardex_filtrado(request)
    if artesano_id:
        kardex = kardex.filter(producto__artesano_id=artesano_id)
    for k in kardex:
        ws2.append([
            k.producto.nombre, k.tipo, k.subtipo or '—',
            k.cantidad, k.stock_resultante,
            k.origen, k.pedido_ref or '—',
            str(k.fecha), k.nota or '—',
        ])
    estilo_excel(ws2, headers2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': 'attachment; filename="inventario.xlsx"'})


@api_view(['GET'])
def reporte_inventario_pdf(request):
    artesano_id = request.query_params.get('artesano')
    productos = Producto.objects.filter(artesano_id=artesano_id) if artesano_id else Producto.objects.all()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []

    # Tabla 1: stock actual
    headers = ['Producto', 'Código', 'Stock actual', 'Stock mín.', 'Stock máx.', 'Estado']
    data = [[
        p.nombre, p.codigo_barra or '—',
        str(p.cantidad), str(p.stock_minimo), str(p.stock_maximo),
        p.estado_stock,
    ] for p in productos]
    estilo_pdf(elements, 'Reporte de Inventario — Pakari Shop', headers, data)

    # Tabla 2: resumen entradas/salidas por producto
    elements.append(Spacer(1, 24))
    headers2 = ['Producto', 'Total entradas', 'Total salidas', 'Neto']
    data2 = []
    for p in productos:
        entradas = Kardex.objects.filter(producto=p, tipo='Entrada').aggregate(t=Sum('cantidad'))['t'] or 0
        salidas  = Kardex.objects.filter(producto=p, tipo='Salida').aggregate(t=Sum('cantidad'))['t'] or 0
        data2.append([p.nombre, f'+{entradas}', f'-{salidas}', str(entradas - salidas)])
    estilo_pdf(elements, 'Resumen de movimientos por producto', headers2, data2)

    doc.build(elements)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf',
                        headers={'Content-Disposition': 'attachment; filename="inventario.pdf"'})


# ── Reportes Kardex ───────────────────────────────────────────────────────────
@api_view(['GET'])
def reporte_kardex_excel(request):
    kardex = _kardex_filtrado(request)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Movimientos'
    headers = ['Producto', 'Tipo', 'Subtipo', 'Cantidad', 'Stock resultante', 'Precio unit.', 'Origen', 'Pedido ref.', 'Fecha', 'Nota', 'Registrado por']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for k in kardex:
        ws.append([
            k.producto.nombre, k.tipo, k.subtipo or '—',
            k.cantidad, k.stock_resultante,
            float(k.precio_unitario) if k.precio_unitario else '—',
            k.origen, k.pedido_ref or '—',
            str(k.fecha), k.nota or '—', k.creado_por,
        ])
    estilo_excel(ws, headers)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': 'attachment; filename="kardex.xlsx"'})


@api_view(['GET'])
def reporte_kardex_pdf(request):
    kardex = _kardex_filtrado(request)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    headers = ['Producto', 'Tipo', 'Cantidad', 'Stock result.', 'Origen', 'Pedido', 'Fecha', 'Nota']
    data = [[
        k.producto.nombre, k.tipo, str(k.cantidad),
        str(k.stock_resultante), k.origen,
        k.pedido_ref or '—', str(k.fecha), k.nota or '—',
    ] for k in kardex]
    estilo_pdf(elements, 'Historial de Movimientos — Pakari Shop', headers, data)
    doc.build(elements)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf',
                        headers={'Content-Disposition': 'attachment; filename="kardex.pdf"'})


# ── Reportes Contable ─────────────────────────────────────────────────────────
@api_view(['GET'])
def reporte_contable_excel(request):
    artesano_id = request.query_params.get('artesano')
    productos = Producto.objects.filter(artesano_id=artesano_id) if artesano_id else Producto.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Contable'
    headers = ['Producto', 'Precio neto', 'IVA (%)', 'Precio con IVA', 'Descuento', 'Precio final (PVP)', 'Stock', 'Valor inventario (neto)']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for p in productos:
        ws.append([
            p.nombre,
            float(p.precio_neto),
            p.iva,
            round(p.precio_con_iva, 2),
            f'{p.valor_descuento}%' if p.descuento else 'No',
            round(p.precio_final, 2),
            p.cantidad,
            round(float(p.precio_neto) * p.cantidad, 2),  # CORREGIDO: precio_neto * stock
        ])
    estilo_excel(ws, headers)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': 'attachment; filename="contable.xlsx"'})


@api_view(['GET'])
def reporte_contable_pdf(request):
    artesano_id = request.query_params.get('artesano')
    productos = Producto.objects.filter(artesano_id=artesano_id) if artesano_id else Producto.objects.all()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    headers = ['Producto', 'Precio neto', 'IVA', 'Precio con IVA', 'Precio final (PVP)', 'Stock', 'Valor inventario (neto)']
    data = [[
        p.nombre,
        f'${float(p.precio_neto):,.0f}',
        f'{p.iva}%',
        f'${p.precio_con_iva:,.0f}',
        f'${p.precio_final:,.0f}',
        str(p.cantidad),
        f'${float(p.precio_neto) * p.cantidad:,.0f}',  # CORREGIDO: precio_neto * stock
    ] for p in productos]
    estilo_pdf(elements, 'Reporte Contable — Pakari Shop', headers, data)
    doc.build(elements)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf',
                        headers={'Content-Disposition': 'attachment; filename="contable.pdf"'})


# ── Notificaciones ────────────────────────────────────────────────────────────
class NotificacionViewSet(viewsets.ModelViewSet):
    queryset = Notificacion.objects.all()
    serializer_class = NotificacionSerializer
    permission_classes = [AllowAny]

    @action(detail=False, methods=['patch'], url_path='leer-todas')
    def leer_todas(self, request):
        Notificacion.objects.filter(leida=False).update(leida=True)
        return Response({'ok': True})

    @action(detail=True, methods=['patch'], url_path='leer')
    def leer(self, request, pk=None):
        notificacion = self.get_object()
        notificacion.leida = True
        notificacion.save()
        return Response({'ok': True})

    @action(detail=False, methods=['get'], url_path='no-leidas')
    def no_leidas(self, request):
        count = Notificacion.objects.filter(leida=False).count()
        return Response({'count': count})